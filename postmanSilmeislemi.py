import requests
from openpyxl import Workbook,load_workbook



wb = load_workbook("C:/Users/SaatTeknoloji/Desktop/Kitap1.xlsx")
ws = wb.active
 

columns = ws["B"]




"""
#tek sütunu yazmak istersen bu şekilde yazman yeterli
for data in columns:
    
    url='http://10.98.228.146:8090/planprofiles/'+'{}'.format(data.value)+'/deactivate'
    
"""

for data in range(1,14): # birden fazla sütunları almak istersen range daha mantıklı
    
    url='http://10.98.225.178:8090/catchups/'+'{}'.format(ws['B{}'.format(data)].value)

    header={}
    payload={}
    
    response = requests.request("DELETE", url, headers=header, data=payload)
    
    print(url)
    
    print(response.text)
      